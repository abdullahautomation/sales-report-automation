import pandas as pd
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, Border, PatternFill, Side, Border
from openpyxl.utils import get_column_letter
def start(file_name, output_file_name):
    file_name = f"{file_name}.xlsx"
    output_file_name= f"{output_file_name}.xlsx"
    try:
        df = pd.read_excel(file_name)
    except Exception as e:
        print(f"Error: {e}")
        return
    # print("----Raw_Data----")
    # print(df)
    df = df.dropna(subset=["Sale_Amount"])
    df["Salesperson"] = df["Salesperson"].fillna("Unknown")
    df["Product"] = df["Product"].fillna("Unknown")
    df["Region"] = df["Region"].fillna("Unknown")
    df["Product"] = df["Product"].fillna("Unknown")
    df["Order_ID"] = df["Order_ID"].fillna(0)
    df["Salesperson"] = df["Salesperson"].str.strip().str.title()
    # print("----Cleaned_Data----")
    # print(df)
    
 

    monthly = df.groupby("Month")["Sale_Amount"].sum().reset_index()
    monthly.columns = ["Month", "Total_Sales"]
    product = df.groupby("Product")["Sale_Amount"].sum().reset_index()
    product.columns = ["Product", "Total_Sales"]
    try:
        with pd.ExcelWriter(output_file_name, engine="openpyxl") as file:
            df.to_excel(file, sheet_name="Cleaned_Data", index=False)
            monthly.to_excel(file, sheet_name="Monthly Summary", index=False)
            product.to_excel(file, sheet_name="Product Summary", index=False)
    except Exception as c:
        print(f"Error : {c}")
        return
    workbook = openpyxl.load_workbook(output_file_name)
    sheet1 = workbook.active
    sheet2 = workbook['Monthly Summary']
    sheet3 = workbook['Product Summary']
    for cell in sheet1[1]:
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="FFFF00"
        )
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for cell in sheet2[1]:
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="FFFF00"
        )
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for cell in sheet3[1]:
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="FFFF00"
        )
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet1.column_dimensions['A'].width = 10
    sheet1.column_dimensions['B'].width = 15
    sheet1.column_dimensions['E'].width = 15
    sheet1.column_dimensions['C'].width = 11
    sheet1.column_dimensions['D'].width = 11
    for col in sheet1.iter_cols(min_row=2, min_col=1):
        for cell in col:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in sheet2.iter_cols(min_row=2, min_col=1):
        for cell in col:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in sheet3.iter_cols(min_row=2, min_col=1):
        for cell in col:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for i in range(1, 3):
        column_letter = get_column_letter(i)
        sheet2.column_dimensions[column_letter].width = 14
        sheet3.column_dimensions[column_letter].width = 14
    thin = Side(style="thin")
    brthin = Border(top=thin, bottom=thin, left=thin, right=thin)
    thick = Side(style="thick")
    brthick = Border(top=thick, bottom=thick, left=thick, right=thick)
    for cell in sheet1[1]:
        cell.border = brthick
    for row in sheet1.iter_rows(min_row=2,min_col=1):
        for cell in row:
            cell.border=brthin
            cell.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    for cell in sheet2[1]:
        cell.border = brthick
    for row in sheet2.iter_rows(min_row=2,min_col=1):
        for cell in row:
            cell.border=brthin
            cell.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    for cell in sheet3[1]:
        cell.border = brthick
    for row in sheet3.iter_rows(min_row=2,min_col=1):
        for cell in row:
            cell.border=brthin
            cell.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    chart2 = BarChart()
    chart2.title = "Monthly Sales"
    chart2.y_axis.title = "Sales Amount"
    chart2.x_axis.title = "Month"
    month_rows = len(monthly) + 1
    sales2 = Reference(sheet2, min_row=1,max_row= month_rows, min_col=2, max_col=2)
    months = Reference(sheet2, min_row=2,max_row= month_rows, min_col=1, max_col=1)
    chart2.add_data(sales2, titles_from_data=True)
    chart2.set_categories(months)
    chart_row = len(monthly) + 3
    sheet2.add_chart(chart2,f'A{chart_row}')
    chart = BarChart()
    chart.title = "Product Summary"
    chart.x_axis.title = "Product"
    chart.y_axis.title = "Sales Amount"
    product_rows = len(product) + 1
    sales = Reference(sheet3, min_row=1, max_row= product_rows , min_col=2, max_col=2)
    product = Reference(sheet3, min_row=2, max_row= product_rows , min_col=1, max_col=1)
    chart.add_data(sales, titles_from_data=True)
    chart.set_categories(product)
    chart_row = len(product) + 3
    sheet3.add_chart(chart, f'A{chart_row}')
    print(monthly)

    workbook.save(output_file_name)
    workbook.close()
    print("Done!")
if __name__ == "__main__":
#     a = str(input("Enter the file name you want to clean: "))
#     b = str(input("Enter the output file name: "))
    start("raw_sales_data.xlsx", "output.xlsx")
